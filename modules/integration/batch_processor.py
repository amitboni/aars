"""modules/integration/batch_processor.py — CSV/Excel file processing.

Reads uploaded files, applies field mapping, routes rows to the appropriate
adapter, and collects success/failure counts.
"""
from __future__ import annotations

import csv
import io
import logging
import uuid

from sqlalchemy.ext.asyncio import AsyncSession

from modules.integration.base import IntegrationAdapter
from modules.integration.commission_adapter import CommissionAdapter
from modules.integration.field_mapping import apply_field_mapping, get_default_mapping
from modules.integration.lms_adapter import LMSLicenseAdapter, LMSTrainingAdapter
from modules.integration.pas_adapter import PASAgentAdapter, PASPolicyAdapter
from modules.integration.schemas import UploadResponse
from modules.signal import service as signal_service

logger = logging.getLogger(__name__)


def _detect_data_type(row: dict, source_system: str) -> str:
    """Detect the data type of a row based on its fields and source system."""
    if source_system == "pas":
        if row.get("policy_number"):
            return "policy"
        return "agent"
    elif source_system == "lms":
        if row.get("training_name"):
            return "training"
        if row.get("license_number"):
            return "license"
        return "training"
    elif source_system == "commission":
        return "commission"
    return "unknown"


def _get_adapter(source_system: str, data_type: str) -> IntegrationAdapter:
    """Return the appropriate adapter for the source system and data type."""
    adapters: dict[str, dict[str, IntegrationAdapter]] = {
        "pas": {
            "agent": PASAgentAdapter(),
            "policy": PASPolicyAdapter(),
        },
        "lms": {
            "training": LMSTrainingAdapter(),
            "license": LMSLicenseAdapter(),
        },
        "commission": {
            "commission": CommissionAdapter(),
        },
    }
    system_adapters = adapters.get(source_system, {})
    adapter = system_adapters.get(data_type)
    if adapter is None:
        raise ValueError(f"No adapter for {source_system}/{data_type}")
    return adapter


def _read_csv(content: bytes) -> list[dict]:
    """Read CSV content into list of dicts."""
    text = content.decode("utf-8-sig")  # Handle BOM
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def _read_excel(content: bytes) -> list[dict]:
    """Read Excel content into list of dicts using openpyxl."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return []

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = val if val is not None else ""
        result.append(row_dict)
    return result


async def process_file(
    content: bytes,
    file_name: str,
    tenant_id: uuid.UUID,
    source_system: str,
    field_mapping: dict[str, str] | None,
    db: AsyncSession,
) -> UploadResponse:
    """Process an uploaded CSV/Excel file.

    For each row:
    1. Apply field mapping
    2. Detect data type and route to appropriate adapter
    3. Adapter validates and returns signals
    4. Emit signals with idempotency keys
    5. Track success/failure counts
    """
    # Read file
    ext = file_name.rsplit(".", 1)[-1].lower() if "." in file_name else "csv"
    if ext in ("xlsx", "xls"):
        rows = _read_excel(content)
    else:
        rows = _read_csv(content)

    # Build effective mapping: tenant overrides + defaults
    effective_mapping = get_default_mapping(source_system)
    if field_mapping:
        effective_mapping.update(field_mapping)

    total = len(rows)
    processed = 0
    failed = 0
    errors: list[dict] = []

    for row_idx, raw_row in enumerate(rows, start=1):
        try:
            # Apply field mapping
            mapped_row = apply_field_mapping(raw_row, effective_mapping)

            # Detect data type and get adapter
            data_type = _detect_data_type(mapped_row, source_system)
            adapter = _get_adapter(source_system, data_type)

            # Validate
            is_valid, validation_errors = adapter.validate_row(mapped_row)
            if not is_valid:
                failed += 1
                errors.append({
                    "row": row_idx,
                    "errors": validation_errors,
                    "data": {k: str(v)[:100] for k, v in mapped_row.items()},
                })
                continue

            # Process row → get signals
            signals = await adapter.process_row(mapped_row, tenant_id, db)

            # Emit each signal
            for signal_data in signals:
                await signal_service.emit_signal(db, tenant_id, signal_data)

            processed += 1

        except Exception as exc:
            failed += 1
            errors.append({
                "row": row_idx,
                "errors": [str(exc)],
                "data": {k: str(v)[:100] for k, v in raw_row.items()},
            })
            logger.warning("Error processing row %d: %s", row_idx, exc)

    # Generate a job_id for tracking
    job_id = uuid.uuid4()

    return UploadResponse(
        job_id=job_id,
        status="completed" if failed == 0 else "completed_with_errors",
        total_records=total,
        processed=processed,
        failed=failed,
        errors=errors[:50],  # Limit error details to first 50
    )
